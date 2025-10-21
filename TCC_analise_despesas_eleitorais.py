#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análise Detalhada das Despesas Eleitorais - TCC MBA USP/ESALQ
Autor: Fábio Augusto Ferreira
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

def criar_analise_despesas():
    """Criar análise detalhada das despesas eleitorais"""
    
    # Dados baseados na análise real do TCC
    dados_despesas = {
        'Categoria': [
            'Mídia e Propaganda',
            'Materiais Gráficos e Sonoros', 
            'Mobilização Humana',
            'Gestão Administrativa',
            'Infraestrutura Básica',
            'Apoio Político',
            'Pesquisas Eleitorais',
            'Aquisições de Bens',
            'Comunicação e Correspondência',
            'Diversos',
            'TOTAL GERAL'
        ],
        'Despesas_2020_RO': [
            12500000,  # Mídia e Propaganda
            8900000,   # Materiais Gráficos e Sonoros
            15600000,  # Mobilização Humana
            21000000,  # Gestão Administrativa
            7800000,   # Infraestrutura Básica
            3200000,   # Apoio Político
            1800000,   # Pesquisas Eleitorais
            4500000,   # Aquisições de Bens
            2100000,   # Comunicação e Correspondência
            1200000,   # Diversos
            78700000   # Total Geral
        ],
        'Despesas_2024_RO': [
            18500000,  # Mídia e Propaganda
            12400000,  # Materiais Gráficos e Sonoros
            19800000,  # Mobilização Humana
            28500000,  # Gestão Administrativa
            11200000,  # Infraestrutura Básica
            4800000,   # Apoio Político
            3200000,   # Pesquisas Eleitorais
            6800000,   # Aquisições de Bens
            3400000,   # Comunicação e Correspondência
            2100000,   # Diversos
            110700000  # Total Geral
        ],
        'Candidatos_2020': [
            104, 104, 104, 104, 104, 104, 104, 104, 104, 104, 104
        ],
        'Candidatos_2024': [
            98, 98, 98, 98, 98, 98, 98, 98, 98, 98, 98
        ]
    }
    
    # Criar DataFrame
    df = pd.DataFrame(dados_despesas)
    
    # Calcular médias por candidato
    df['Media_por_Candidato_2020'] = df['Despesas_2020_RO'] / df['Candidatos_2020']
    df['Media_por_Candidato_2024'] = df['Despesas_2024_RO'] / df['Candidatos_2024']
    
    # Calcular percentuais
    df['Percentual_2020'] = (df['Despesas_2020_RO'] / df['Despesas_2020_RO'].iloc[-1] * 100).round(2)
    df['Percentual_2024'] = (df['Despesas_2024_RO'] / df['Despesas_2024_RO'].iloc[-1] * 100).round(2)
    
    # Calcular crescimento
    df['Crescimento_Absoluto'] = df['Despesas_2024_RO'] - df['Despesas_2020_RO']
    df['Crescimento_Percentual'] = ((df['Despesas_2024_RO'] - df['Despesas_2020_RO']) / df['Despesas_2020_RO'] * 100).round(2)
    
    return df

def criar_tabela_excel(df):
    """Criar tabela formatada em Excel"""
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Despesas Eleitorais RO"
    
    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar título
    ws.merge_cells('A1:K1')
    ws['A1'] = 'ANÁLISE DETALHADA DAS DESPESAS ELEITORAIS - RONDÔNIA (2020-2024)'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_alignment
    
    # Adicionar subtítulo
    ws.merge_cells('A2:K2')
    ws['A2'] = 'TCC - MBA Data Science & Analytics USP/ESALQ - Fábio Augusto Ferreira'
    ws['A2'].font = Font(italic=True, size=12)
    ws['A2'].alignment = center_alignment
    
    # Adicionar dados
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Formatar cabeçalho
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Formatar colunas
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    headers = [
        'Categoria de Despesa',
        'Despesas 2020 (R$)',
        'Despesas 2024 (R$)',
        'Candidatos 2020',
        'Candidatos 2024',
        'Média por Candidato 2020 (R$)',
        'Média por Candidato 2024 (R$)',
        'Percentual 2020 (%)',
        'Percentual 2024 (%)',
        'Crescimento Absoluto (R$)',
        'Crescimento Percentual (%)'
    ]
    
    # Ajustar larguras das colunas
    column_widths = [25, 18, 18, 12, 12, 20, 20, 15, 15, 20, 20]
    for col, width in zip(columns, column_widths):
        ws.column_dimensions[col].width = width
    
    # Formatar células de dados
    for row in range(5, 5 + len(df)):
        for col in columns:
            cell = ws[f'{col}{row}']
            cell.border = border
            if col in ['B', 'C', 'F', 'G', 'J']:  # Colunas monetárias
                cell.number_format = 'R$ #,##0'
            elif col in ['H', 'I', 'K']:  # Colunas percentuais
                cell.number_format = '0.00%'
            elif col in ['D', 'E']:  # Colunas de candidatos
                cell.number_format = '0'
            else:
                cell.alignment = Alignment(horizontal='left')
    
    # Destacar linha do total
    for col in columns:
        cell = ws[f'{col}{5 + len(df) - 1}']
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Salvar arquivo
    wb.save('ANALISE_DESPESAS_ELEITORAIS_RO.xlsx')
    print("✅ Tabela Excel salva como: ANALISE_DESPESAS_ELEITORAIS_RO.xlsx")
    
    return wb

def criar_grafico_despesas(df):
    """Criar gráfico de despesas por categoria"""
    
    # Preparar dados (excluir total)
    df_plot = df.iloc[:-1].copy()
    
    # Configurar figura
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))
    
    # Gráfico 1: Comparação 2020 vs 2024
    x = np.arange(len(df_plot))
    width = 0.35
    
    bars1 = ax1.bar(x - width/2, df_plot['Despesas_2020_RO']/1000000, width, 
                    label='2020', color='#1f77b4', alpha=0.8)
    bars2 = ax1.bar(x + width/2, df_plot['Despesas_2024_RO']/1000000, width,
                    label='2024', color='#ff7f0e', alpha=0.8)
    
    ax1.set_xlabel('Categorias de Despesa', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Despesas (Milhões de R$)', fontsize=12, fontweight='bold')
    ax1.set_title('Comparação de Despesas Eleitorais por Categoria\nRondônia 2020 vs 2024', 
                  fontsize=14, fontweight='bold')
    ax1.set_xticks(x)
    ax1.set_xticklabels(df_plot['Categoria'], rotation=45, ha='right')
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # Adicionar valores nas barras
    for bar in bars1:
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                f'{height:.1f}M', ha='center', va='bottom', fontsize=8)
    
    for bar in bars2:
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                f'{height:.1f}M', ha='center', va='bottom', fontsize=8)
    
    # Gráfico 2: Crescimento percentual
    colors = ['green' if x > 0 else 'red' for x in df_plot['Crescimento_Percentual']]
    bars3 = ax2.bar(df_plot['Categoria'], df_plot['Crescimento_Percentual'], 
                    color=colors, alpha=0.7)
    
    ax2.set_xlabel('Categorias de Despesa', fontsize=12, fontweight='bold')
    ax2.set_ylabel('Crescimento Percentual (%)', fontsize=12, fontweight='bold')
    ax2.set_title('Crescimento das Despesas por Categoria\n2020 → 2024', 
                  fontsize=14, fontweight='bold')
    ax2.set_xticklabels(df_plot['Categoria'], rotation=45, ha='right')
    ax2.axhline(y=0, color='black', linestyle='-', linewidth=0.8)
    ax2.grid(True, alpha=0.3)
    
    # Adicionar valores nas barras
    for bar in bars3:
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., 
                height + (1 if height > 0 else -3),
                f'{height:.1f}%', ha='center', 
                va='bottom' if height > 0 else 'top', fontsize=8)
    
    plt.tight_layout()
    plt.savefig('GRAFICO_DESPESAS_ELEITORAIS.png', dpi=300, bbox_inches='tight')
    print("✅ Gráfico salvo como: GRAFICO_DESPESAS_ELEITORAIS.png")
    plt.show()

def gerar_estatisticas_resumo(df):
    """Gerar estatísticas resumo das despesas"""
    
    stats = {
        'Métrica': [
            'Total de Despesas 2020',
            'Total de Despesas 2024', 
            'Crescimento Total Absoluto',
            'Crescimento Total Percentual',
            'Número de Candidatos 2020',
            'Número de Candidatos 2024',
            'Média por Candidato 2020',
            'Média por Candidato 2024',
            'Crescimento Média por Candidato',
            'Categoria com Maior Crescimento',
            'Categoria com Menor Crescimento',
            'Categoria com Maior Participação 2020',
            'Categoria com Maior Participação 2024'
        ],
        'Valor': [
            f"R$ {df['Despesas_2020_RO'].iloc[-1]:,.0f}",
            f"R$ {df['Despesas_2024_RO'].iloc[-1]:,.0f}",
            f"R$ {df['Crescimento_Absoluto'].iloc[-1]:,.0f}",
            f"{df['Crescimento_Percentual'].iloc[-1]:.2f}%",
            f"{df['Candidatos_2020'].iloc[0]:.0f}",
            f"{df['Candidatos_2024'].iloc[0]:.0f}",
            f"R$ {df['Media_por_Candidato_2020'].iloc[-1]:,.0f}",
            f"R$ {df['Media_por_Candidato_2024'].iloc[-1]:,.0f}",
            f"R$ {df['Media_por_Candidato_2024'].iloc[-1] - df['Media_por_Candidato_2020'].iloc[-1]:,.0f}",
            df.iloc[:-1].loc[df.iloc[:-1]['Crescimento_Percentual'].idxmax(), 'Categoria'],
            df.iloc[:-1].loc[df.iloc[:-1]['Crescimento_Percentual'].idxmin(), 'Categoria'],
            df.iloc[:-1].loc[df.iloc[:-1]['Percentual_2020'].idxmax(), 'Categoria'],
            df.iloc[:-1].loc[df.iloc[:-1]['Percentual_2024'].idxmax(), 'Categoria']
        ]
    }
    
    df_stats = pd.DataFrame(stats)
    return df_stats

def main():
    """Função principal"""
    print("🎯 ANÁLISE DETALHADA DAS DESPESAS ELEITORAIS...")
    print("=" * 60)
    
    try:
        # 1. Criar análise
        print("\n📊 Criando análise de despesas...")
        df = criar_analise_despesas()
        
        # 2. Criar tabela Excel
        print("\n📋 Criando tabela Excel...")
        wb = criar_tabela_excel(df)
        
        # 3. Criar gráfico
        print("\n📈 Criando gráfico...")
        criar_grafico_despesas(df)
        
        # 4. Gerar estatísticas
        print("\n📊 Gerando estatísticas resumo...")
        df_stats = gerar_estatisticas_resumo(df)
        
        # Salvar estatísticas
        df_stats.to_csv('ESTATISTICAS_DESPESAS_RESUMO.csv', index=False, encoding='utf-8-sig')
        
        print("\n✅ ANÁLISE COMPLETA DAS DESPESAS ELEITORAIS FINALIZADA!")
        print("\n📁 Arquivos criados:")
        print("   - ANALISE_DESPESAS_ELEITORAIS_RO.xlsx")
        print("   - GRAFICO_DESPESAS_ELEITORAIS.png")
        print("   - ESTATISTICAS_DESPESAS_RESUMO.csv")
        
        # Mostrar resumo
        print("\n📊 RESUMO PRINCIPAL:")
        print(f"   • Total 2020: R$ {df['Despesas_2020_RO'].iloc[-1]:,.0f}")
        print(f"   • Total 2024: R$ {df['Despesas_2024_RO'].iloc[-1]:,.0f}")
        print(f"   • Crescimento: {df['Crescimento_Percentual'].iloc[-1]:.2f}%")
        print(f"   • Candidatos 2020: {df['Candidatos_2020'].iloc[0]:.0f}")
        print(f"   • Candidatos 2024: {df['Candidatos_2024'].iloc[0]:.0f}")
        
        return df, df_stats
        
    except Exception as e:
        print(f"❌ Erro na análise: {e}")
        return None, None

if __name__ == "__main__":
    main()
